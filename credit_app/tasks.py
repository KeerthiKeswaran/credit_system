from celery import shared_task
from .models import Customer, Loan
import openpyxl
from datetime import datetime
import os

@shared_task
def load_initial_data():
    base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    customer_path = os.path.join(base_path, 'customer_data.xlsx')
    loan_path = os.path.join(base_path, 'loan_data.xlsx')

    # Load Customers
    wb = openpyxl.load_workbook(customer_path)
    sheet = wb.active
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        Customer.objects.update_or_create(
            phone_number=row[3],
            defaults={
                'first_name': row[1],
                'last_name': row[2],
                'monthly_salary': row[4],
                'approved_limit': row[5],
                'current_debt': row[6]
            }
        )

    # Load Loans
    wb = openpyxl.load_workbook(loan_path)
    sheet = wb.active
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        customer_id = row[0]
        customer = Customer.objects.filter(id=customer_id).first()
        if customer:
            Loan.objects.update_or_create(
                loan_id=row[1],
                defaults={
                    'customer': customer,
                    'loan_amount': row[2],
                    'tenure': row[3],
                    'interest_rate': row[4],
                    'monthly_installment': row[5],
                    'emis_paid_on_time': row[6],
                    'start_date': datetime.strptime(str(row[7]), '%Y-%m-%d').date(),
                    'end_date': datetime.strptime(str(row[8]), '%Y-%m-%d').date()
                }
            )
